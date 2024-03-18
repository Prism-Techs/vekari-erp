from io import BytesIO
import os

from django.http import FileResponse, HttpResponse
from qc_reports.models import QcReportsMaster
from qc_reports.serializers import QCReportSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.permissions import IsAuthenticated , IsAdminUser,AllowAny
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment,Font
from django.conf import settings
from datetime import datetime
from openpyxl.utils import get_column_letter
from authuser.permission import YourPermission 

from rest_framework import status

class ExcelQCReportView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,YourPermission]
    group_permissions = ['quality']   

    def get(self,*args, **kwargs):
        qcr_id=kwargs.get('qcr_id')
        if qcr_id:
            instance=QcReportsMaster.objects.filter(qcr_id=qcr_id)
            if instance.exists():
                instance=instance.first()
                serializer= QCReportSerializer(instance=instance)
                logo_path=os.path.join(settings.BASE_DIR, "media","new_vekaria_logo.png")
                excel_path=qc_excel_report(serializer.data,'Vekaria Engineering Works Private Limited',logo_path)
                excel_path=excel_path.replace('\\','/')
                instance.qc_report_excel=excel_path
                instance.save()
                
                return  Response({"qc_report_excel":excel_path},status=status.HTTP_200_OK)

            return Response({"message":"QC is not found."},status=status.HTTP_404_NOT_FOUND)
        return Response({"message":"QC Id is required."},status=status.HTTP_400_BAD_REQUEST)
  


def qc_excel_report(data, title, logo_path):
    # Ensure the directory exists
    # print(data)
    directory_path = os.path.join(settings.BASE_DIR, "media", "qc_report")
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    # start title 
    # Insert title in the second column of the first row
    ws.merge_cells('D1:K1')
    ws['D1'] = title
    ws['D1'].alignment =  Alignment(horizontal='center', vertical='center')
    ws['D1'].font = Font(size=12, bold=True) 
    # end title 
    # qc no and date 
    
    ws.merge_cells('L1:O1')
    created_date_str = data.get('created_date')
    if created_date_str:
        # Convert the string to a datetime object
        created_date = datetime.fromisoformat(created_date_str.rstrip('Z'))
        # Format the date as needed and combine with qcr_no
        ws['L1'] = f"{data.get('qcr_no')}/{created_date:%d/%m/%Y}"
    else:
        ws['L1'] = data.get('qcr_no')
        
    ws['L1'].alignment =  Alignment(horizontal='center', vertical='center')
    ws['L1'].font = Font(size=12, bold=True) 
    # Start Add the logo image

    img = Image(logo_path)
    img.width = 206
    img.height = 45
    # Place the image in the first cell (adjust as needed)
    ws.add_image(img, 'A1')
    ws.merge_cells('A1:C1')
    
    # ws.column_dimensions['A'].width = 30  # Adjust col  as needed
    ws.row_dimensions[1].height = 35   # Adjust row as needed to fit the image
    ws.column_dimensions['A'].width = 11 
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10 

    # end logo 


    # row 2 
    # Quality Inspection Report	title																				
    
    ws.merge_cells('A2:C2')
    ws.merge_cells('D2:O2')
    
    ws['D2'] = "Quality Inspection Report"
    ws['D2'].alignment =  Alignment(horizontal='center', vertical='center')
    ws['D2'].font = Font(size=10, bold=True) 

    # end row 2

    # QC Main Detail 
    received_date=data.get('received_date')
    receiv_date=''
    if received_date:
        received_date_z = datetime.fromisoformat(created_date_str.rstrip('Z'))
        receiv_date= f"{received_date_z:%d-%m-%Y}"
    qc_main={
        "Supplier":f"{data.get('vendor_name','')}, {data.get('vendor_address','')} ,{data.get('vendor_city','')}",
        # "Supplier City":data.get('vendor_city',''),
        "Challan No":'',
        'Part No': data.get('product_part_no',''),
        'Part Name': data.get('product_part_name',''),
        "Part Desc":f"{data.get('product_descp','')}",
        "Quantity":f"{data.get('qty','')} NOS",
        "Location":f"Tub/Rack: {data.get('store_loc_rack_no')} & {data.get('store_loc_tub_no')}",
        "Received Date":receiv_date,
        "Received Qty":f"{data.get('rejected_qty','')} NOS",
        "Inspection Qty":f"{data.get('inspt_qty','')} NOS"
    }
    row_no = 3
    for key, value in qc_main.items():
        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=2)
        ws[f'A{row_no}'] = key
        ws[f'A{row_no}'].alignment = Alignment(horizontal='left')

        # Set separator in column C
        ws[f'C{row_no}'] = ":-"

        # Set value in column D and merge cells from D to O
        ws[f'D{row_no}'] = value
        ws.merge_cells(start_row=row_no, start_column=4, end_row=row_no, end_column=15)
        ws[f'D{row_no}'].alignment = Alignment(horizontal='left')

        row_no += 1  # Move to the next row
    
    # End QC Main Detail 


    # Start QC Extra Detail 

    # Start with row number 14 for the headers

    # Set up headers
    row_no = 14
    headers = ['Sr. No.', 'Parameter Name', 'Inspection Name', 'Nominal Dimension', 'Tolerance Negative', 'Tolerance Positive','Tolerance Other']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=row_no, column=col, value=header)
        ws.merge_cells(start_row=row_no, start_column=col, end_row=row_no + 1, end_column=col)
        cell = ws.cell(row=row_no, column=col)
        cell.font = Font(size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = len(header) + 2
     # Adjust column widths based on headers
 
    # Add headers for "Measured Dimensions"
    row_no = 14
    start_col_index = 8  # Column 'H'
    qty = data.get('qty', 0)

    # Calculate end column index based on quantity
    end_col_index = start_col_index + qty - 1

    # Merge cells for "Measured Dimensions" header
    start_col_letter = get_column_letter(start_col_index)
    end_col_letter = get_column_letter(end_col_index)
    merge_range = f'{start_col_letter}{row_no}:{end_col_letter}{row_no}'
    ws.merge_cells(merge_range)

    # Set the header value, font, and alignment
    cell = ws.cell(row=row_no, column=start_col_index)
    cell.value = "Measured Dimensions"
    cell.font = Font(size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths (optional) and row height for better visibility
    for col in range(start_col_index, end_col_index + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.row_dimensions[row_no].height = 20

															
    # Starting row for "Measured Dimensions" header is 15
    row_no = 15

    # Begin at column 'H', which is the 8th column

    for i in range(qty):
        
        cell = ws.cell(row=row_no, column=start_col_index + i)
        cell.value = i+1
        cell.font = Font(size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')


    # Adjust column widths based on headers
    for col, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col)].width = len(header) + 2

#  End  headers for "Measured Dimensions"
    # data for "Measured Dimensions"
    row_no = 16  # Starting row for measured dimensions
    start_col_index = 8  # Starting column (H)



    # Append QC detail rows
    qcr_details = data.get('qcr_details', [])
    print(f"Number of detail entries: {len(qcr_details)}")
    row_no = 16


    for idx, qcd in enumerate(qcr_details, start=1):
        # Extract the data
        parameter_name = qcd.get('parameter_name', '')
        nominal_dimension = qcd.get('nominal_dimension', '')
        tolerance_negative = qcd.get('tolerance_negative', '')
        tolerance_positive = qcd.get('tolerance_positive', '')
        tolerance_other =qcd.get('tolerance_other', '-')
        inspt_type_name = qcd.get('inspt_type_name', '')

        measured_list = qcd.get('measured_list', [])

        col_index = start_col_index  # Reset column index for each qcr detail
        for measured in measured_list:
            measured_dimensions = measured.get('measured_dimensions', '-')
            cell = ws.cell(row=row_no, column=col_index)
            cell.value = measured_dimensions
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Increase column index for the next measured dimension
            col_index += 1
       
        # Prepare and append the row at the specified row number
        row = [idx, parameter_name, inspt_type_name, nominal_dimension, tolerance_negative, tolerance_positive,tolerance_other]
        for col_num, entry in enumerate(row, start=1):
            ws.cell(row=row_no, column=col_num, value=entry)

        # Increment the row number for the next set of qcr_details
        row_no += 1

        # End QC Extra Detail 

        # Define the path for saving the Excel report
        qcr_no = data.get('qcr_no', 'default_qcr_no')
        excel_path = os.path.join(directory_path, f"{qcr_no}_qc_report.xlsx")
        # Save the workbook to the specified file
    wb.save(excel_path)

    path= os.path.join("media", "qc_report",f"{qcr_no}_qc_report.xlsx")
    return path


